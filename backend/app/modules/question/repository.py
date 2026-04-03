import logging

from fastapi import HTTPException, status
from app.models.question.model import Question
from sqlalchemy import func, select, desc
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession

from .schemas import (
    QuestionCreateRequest,
    QuestionListRequest,
    QuestionListResponse,
    QuestionBulkDeleteRequest,
)
from app.models.user.model import User

logger = logging.getLogger(__name__)


class QuestionRepository:
    async def create_question(
        self, session: AsyncSession, data: QuestionCreateRequest
    ) -> Question:
        # No unique check on text requested, but it's often good practice. 
        # For now, just create it.
        
        new_question = Question(
            subject_id=data.subject_id,
            user_id=data.user_id,
            text=data.text,
            option_a=data.option_a,
            option_b=data.option_b,
            option_c=data.option_c,
            option_d=data.option_d,
        )
        session.add(new_question)

        try:
            await session.commit()
            await session.refresh(new_question)
        except Exception:
            await session.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Database error",
            )
        return new_question

    async def get_question(
        self, session: AsyncSession, question_id: int, current_user: User
    ) -> Question:
        stmt = select(Question).options(
            selectinload(Question.subject),
            selectinload(Question.user),
        ).where(Question.id == question_id)
        result = await session.execute(stmt)
        question = result.scalar_one_or_none()

        if not question:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Question not found"
            )

        # Check ownership for teachers
        is_admin = any(role.name.lower() == "admin" for role in current_user.roles)
        if not is_admin and question.user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, detail="Access denied: you can only access your own questions"
            )

        return question

    async def list_questions(
        self, session: AsyncSession, request: QuestionListRequest, current_user: User
    ) -> QuestionListResponse:
        stmt = select(Question).options(
            selectinload(Question.subject),
            selectinload(Question.user),
        )

        # Check if user is teacher (not admin)
        is_teacher = any(role.name.lower() == "teacher" for role in current_user.roles)
        is_admin = any(role.name.lower() == "admin" for role in current_user.roles)

        if not is_admin and is_teacher:
            # Teachers can only see their own questions
            stmt = stmt.where(Question.user_id == current_user.id)

        if request.text:
            stmt = stmt.where(Question.text.ilike(f"%{request.text}%"))
        
        if request.subject_id:
            stmt = stmt.where(Question.subject_id == request.subject_id)

        if request.user_id:
            stmt = stmt.where(Question.user_id == request.user_id)

        stmt = stmt.order_by(desc(Question.created_at))
        stmt = stmt.offset(request.offset).limit(request.limit)

        result = await session.execute(stmt)
        questions = result.scalars().all()

        count_stmt = select(func.count()).select_from(Question)
        if not is_admin and is_teacher:
            count_stmt = count_stmt.where(Question.user_id == current_user.id)
        if request.text:
            count_stmt = count_stmt.where(Question.text.ilike(f"%{request.text}%"))
        if request.subject_id:
            count_stmt = count_stmt.where(Question.subject_id == request.subject_id)
        if request.user_id:
            count_stmt = count_stmt.where(Question.user_id == request.user_id)

        total_result = await session.execute(count_stmt)
        total = total_result.scalar() or 0

        return QuestionListResponse(
            total=total, page=request.page, limit=request.limit, questions=questions
        )

    async def update_question(
        self, session: AsyncSession, question_id: int, data: QuestionCreateRequest, current_user: User
    ) -> Question:
        stmt = select(Question).options(
            selectinload(Question.subject),
            selectinload(Question.user),
        ).where(Question.id == question_id)
        result = await session.execute(stmt)
        question = result.scalar_one_or_none()

        if not question:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Question not found"
            )

        # Check ownership for teachers
        is_admin = any(role.name.lower() == "admin" for role in current_user.roles)
        if not is_admin and question.user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, detail="Access denied: you can only update your own questions"
            )
        
        question.subject_id = data.subject_id
        question.user_id = data.user_id
        question.text = data.text
        question.option_a = data.option_a
        question.option_b = data.option_b
        question.option_c = data.option_c
        question.option_d = data.option_d

        await session.commit()
        await session.refresh(question)
        return question

    async def delete_question(
        self, session: AsyncSession, question_id: int, current_user: User
    ) -> None:
        stmt = select(Question).where(Question.id == question_id)
        result = await session.execute(stmt)
        question = result.scalar_one_or_none()

        if not question:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND, detail="Question not found"
            )

        # Check ownership for teachers
        is_admin = any(role.name.lower() == "admin" for role in current_user.roles)
        if not is_admin and question.user_id != current_user.id:
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN, detail="Access denied: you can only delete your own questions"
            )

        await session.delete(question)
        await session.commit()

    async def bulk_delete_questions(
        self, session: AsyncSession, data: QuestionBulkDeleteRequest, current_user: User
    ) -> int:
        from sqlalchemy import delete

        # Check ownership for teachers
        is_admin = any(role.name.lower() == "admin" for role in current_user.roles)
        if not is_admin:
            if data.user_id != current_user.id:
                raise HTTPException(
                    status_code=status.HTTP_403_FORBIDDEN, detail="Access denied: you can only delete your own questions"
                )
        
        stmt = delete(Question).where(
            Question.subject_id == data.subject_id,
            Question.user_id == data.user_id
        )
        
        result = await session.execute(stmt)
        await session.commit()
        
        return result.rowcount


    async def upload_image(self, file) -> str:
        import shutil
        import uuid
        import os

        # Generate unique filename
        file_ext = file.filename.split(".")[-1]
        filename = f"{uuid.uuid4()}.{file_ext}"
        
        # Use config for upload dir
        # Ensure dir exists (though ideally create on startup or here)
        os.makedirs(settings.file_url.upload_dir, exist_ok=True)
        file_path = f"{settings.file_url.upload_dir}/{filename}"

        with open(file_path, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        
        # Use config for http url
        return f"{settings.file_url.http}/{filename}"

    async def upload_questions_excel(
        self, session: AsyncSession, file, subject_id: int, user_id: int
    ) -> list[Question]:
        import pandas as pd
        import io

        contents = await file.read()
        df = pd.read_excel(io.BytesIO(contents))
        
        # Verify there are at least 5 columns
        if len(df.columns) < 5:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Excel file must contain at least 5 columns (question, option A, option B, option C, option D)",
            )
        
        questions = []
        for index, row in df.iterrows():
            # If subject_id is in row, use it, else use param
            # If image is in row, use it
            
            # Using positional indices instead of column names
            text = str(row.iloc[0]) if not pd.isna(row.iloc[0]) else ""
            opt_a = str(row.iloc[1]) if not pd.isna(row.iloc[1]) else ""
            opt_b = str(row.iloc[2]) if not pd.isna(row.iloc[2]) else ""
            opt_c = str(row.iloc[3]) if not pd.isna(row.iloc[3]) else ""
            opt_d = str(row.iloc[4]) if not pd.isna(row.iloc[4]) else ""
            
            q_subject_id = subject_id
            if "subject_id" in df.columns and not pd.isna(row["subject_id"]):
                 try:
                    q_subject_id = int(row["subject_id"])
                 except:
                    pass
            
            question = Question(
                subject_id=q_subject_id,
                user_id=user_id,
                text=text,
                option_a=opt_a,
                option_b=opt_b,
                option_c=opt_c,
                option_d=opt_d,
            )
            questions.append(question)
            
        session.add_all(questions)
        
        try:
            await session.commit()
        except Exception:
            await session.rollback()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Database error during bulk upload",
            )

        return questions

    async def download_questions_excel(
        self,
        session: AsyncSession,
        subject_id: int | None = None,
        user_id: int | None = None,
        text: str | None = None,
    ) -> bytes:
        import io
        import re
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        def strip_html(html: str) -> str:
            """Remove HTML tags and return plain text."""
            clean = re.sub(r'<[^>]+>', '', html or '')
            return clean.strip()

        # Query all matching questions (no pagination)
        stmt = select(Question).options(
            selectinload(Question.subject),
            selectinload(Question.user),
        )

        if text:
            stmt = stmt.where(Question.text.ilike(f"%{text}%"))
        if subject_id:
            stmt = stmt.where(Question.subject_id == subject_id)
        if user_id:
            stmt = stmt.where(Question.user_id == user_id)

        stmt = stmt.order_by(desc(Question.created_at))

        result = await session.execute(stmt)
        questions = result.scalars().all()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Savollar"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        headers = ["№", "Savol", "A variant", "B variant", "C variant", "D variant", "Fan", "Foydalanuvchi"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Data rows
        cell_alignment = Alignment(vertical="top", wrap_text=True)
        for row_idx, q in enumerate(questions, 2):
            subject_name = q.subject.name if q.subject else "-"
            username = q.user.username if q.user else "-"

            values = [
                row_idx - 1,
                strip_html(q.text),
                strip_html(q.option_a),
                strip_html(q.option_b),
                strip_html(q.option_c),
                strip_html(q.option_d),
                subject_name,
                username,
            ]
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border

        # Column widths
        ws.column_dimensions["A"].width = 6    # №
        ws.column_dimensions["B"].width = 50   # Savol
        ws.column_dimensions["C"].width = 25   # A
        ws.column_dimensions["D"].width = 25   # B
        ws.column_dimensions["E"].width = 25   # C
        ws.column_dimensions["F"].width = 25   # D
        ws.column_dimensions["G"].width = 20   # Fan
        ws.column_dimensions["H"].width = 18   # Foydalanuvchi

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

get_question_repository = QuestionRepository()
